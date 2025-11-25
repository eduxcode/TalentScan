import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from ..models import JobPosition

class ReportGenerator:
    """Service to generate Excel reports"""

    @staticmethod
    def generate_excel(job_id):
        """
        Generates an Excel report for a specific job.
        
        Args:
            job_id (int): ID of the job position.
            
        Returns:
            Workbook: Openpyxl workbook object.
        """
        try:
            job = JobPosition.objects.get(id=job_id)
        except JobPosition.DoesNotExist:
            return None

        candidates = job.candidates.select_related('analysis').all()
        criteria_list = list(job.criteria.all())

        # Prepare data
        data = []
        for cand in candidates:
            row = {
                'Nome': cand.name,
                'Email': cand.email,
                'Telefone': cand.phone,
                'Arquivo CV': cand.cv_file.name.split('/')[-1] if cand.cv_file else '',
                'Pontuação Total': getattr(cand.analysis, 'total_score', 0) if hasattr(cand, 'analysis') else 0,
                'Resumo': getattr(cand.analysis, 'summary', '') if hasattr(cand, 'analysis') else '',
            }
            
            # Add individual criteria scores
            if hasattr(cand, 'analysis'):
                scores = cand.analysis.data_json.get('scores', {})
                for crit in criteria_list:
                    row[crit.name] = scores.get(crit.name, 0)
            
            data.append(row)

        df = pd.DataFrame(data)

        # Create Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Análise de Currículos"

        # Add headers
        headers = ['Nome', 'Email', 'Telefone', 'Arquivo CV', 'Pontuação Total'] + [c.name for c in criteria_list] + ['Resumo']
        ws.append(headers)

        # Add data
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)

        # Formatting Main Sheet
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = min(adjusted_width, 50)

        # --- Statistics Sheet ---
        if not df.empty:
            ws_stats = wb.create_sheet("Resumo")
            ws_stats.append(["Métrica", "Valor"])
            
            stats = [
                ("Total de Candidatos", len(df)),
                ("Média Geral", df['Pontuação Total'].mean()),
                ("Maior Pontuação", df['Pontuação Total'].max()),
                ("Menor Pontuação", df['Pontuação Total'].min()),
            ]
            
            # Criteria Averages
            for crit in criteria_list:
                if crit.name in df.columns:
                    stats.append((f"Média - {crit.name}", df[crit.name].mean()))

            for stat in stats:
                ws_stats.append(stat)

            # Format Stats Header
            for cell in ws_stats[1]:
                cell.font = header_font
                cell.fill = header_fill

            ws_stats.column_dimensions['A'].width = 30
            ws_stats.column_dimensions['B'].width = 15

        return wb
